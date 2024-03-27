import openpyxl as pxl
import random
import ast

# Open and setup doc
wb = pxl.load_workbook('../doc.xls', data_only=True)
sheets = wb.sheetnames

def read_file(filename, index):
    with open(filename, 'r') as file:
        lines = file.readlines() # Read the file
        if index < len(lines):
            return lines[index].strip()
        else:
            return None  # Index out of range

def read_excel(sheet, cell): # Read the value on a cell, in a specified sheet of the doc
    sheetName = str(sheet)
    ws = wb[sheetName]
    return ws[str(cell)].value

class Weapon:
    def __init__(self, name, hit, crit, miss):
        self.name = name
        # self.rolls = int(rolls)
        self.hit = int(hit)
        self.crit = crit
        self.miss = miss
    
    def display(self):
        return self.name, self.hit, self.crit, self.miss

class Item:
    def __init__(self, name, defense, dodge, regen):
        self.name = name
        self.defense = defense
        self.dodge = dodge
        self.regen = regen
    
    def display(self):
        return self.name, self.defense, self.dodge, self.regen

class Player:
    def __init__(self, level, role, weapon_ID, item_ID, regen):
        self.role = str(role)
        self.level = str(level)
        self.weapon = Weapon( # Using the Weapon class in the Player class here
            name = read_excel("Weapons", str("C" + str(int(weapon_ID) + 1))),
            hit = read_excel("Weapons", str("D" + str(int(weapon_ID) + 1))),
            crit = read_excel("Weapons", str("E" + str(int(item_ID) + 1))),
            miss = read_excel("Weapons", str("F" + str(int(item_ID) + 1)))
        )
        self.item = Item(
            name = read_excel("Items", str("C" + str(int(item_ID) + 1))),
            defense = read_excel("Items", str("D" + str(int(item_ID) + 1))),
            dodge = read_excel("Items", str("E" + str(int(item_ID) + 1))),
            regen = read_excel("Items", str("F" + str(int(item_ID) + 1))) # + role_regen
        )
        self.regen = self.item.regen

    def calculate_damage(self): # Tries to deal damage
        if random.randint(1, 100) <= int(self.weapon.miss):
            return 0
    
        damage = random.gauss(self.weapon.hit, self.weapon.hit/6) # Gauss for damage calculation

        #for _ in range(self.weapon.rolls):
        #    roll_result = random.randint(1, 6)  # Roll a dice
        #    if roll_result >= self.weapon.hit_threshold:
        #        damage += 1
        
        if random.randint(1, 100) <= int(self.weapon.crit):
            return round(damage*2)
        else:
            return round(damage)

    def damaged(self, dmg): # Tries to take damage
        if random.randint(1, 100) >= int(self.item.dodge):
            return True, dmg - int(self.item.defense)/100*dmg
        return False, 0
        
    def display(self):
        return self.role, self.level, self.weapon.display(), self.item.display()

class Enemy:
    def __init__(self, name, hp, hit, miss, defense, dodge, regen, poison):
        self.name = name
        self.hp = hp
        self.hit = hit
        self.miss = miss
        self.defense = defense
        self.dodge = dodge
        self.regen = regen
        self.poison = poison
    
    def calculate_damage(self): # Tries to deal damage
        if random.randint(1, 100) <= int(self.dodge):
            return 0
    
        damage = random.gauss(self.hit, self.hit/6) # Gauss for damage calculation

    def damaged(self, dmg): # Tries to take damage
        if random.randint(1, 100) >= int(self.dodge):
            return True, dmg - int(self.defense)/100*dmg
        return False, 0
        
    def display(self):
        return self.name, self.hp, self.hit, self.miss, self.defense, self.dodge, self.regen, self.poison

def main():
    # Values defined by a tk interface, written in a config.txt file
    config = ast.literal_eval(read_file("config.txt", 0)) # Convert the config file's info to array
    level = ast.literal_eval(read_file("level.txt", 0)) # Read the level info

    player_count = len(config)
    players = [] # Contains player objects

    for i in range(player_count):
        # Add a player
        players.append(
            Player(
            role = config[i][0],
            level = config[i][1],
            weapon_ID = config[i][2],
            item_ID = config[i][3],
            regen = config[i][4]
            )
        )

    Enemy = Enemy(
        name = read_excel("Enemies", str("B" + str(int(level[0]) + 1))),
        hp = read_excel("Enemies", str("C" + str(int(level[0]) + 1))),
        hit = read_excel("Enemies", str("D" + str(int(level[0]) + 1))),
        miss = read_excel("Enemies", str("E" + str(int(level[0]) + 1))),
        defense = read_excel("Enemies", str("F" + str(int(level[0]) + 1))),
        dodge = read_excel("Enemies", str("G" + str(int(level[0]) + 1))),
        regen = read_excel("Enemies", str("H" + str(int(level[0]) + 1))),
        poison = read_excel("Enemies", str("I" + str(int(level[0]) + 1)))
    )

    team_regen = 0
    for i in range(player_count): # Add healer's regen stat to all the team
        if players[i].role == "M" or players[i].role == "D":
            team_regen += int(players[i].regen)
    
    # Debug tests here: -----------------

    print("Team regen : " + str(team_regen) + "\n")

    # Then temporary, just debugging for now
    for i in range(player_count):
        print(players[i].display())

        print("Player " + str(i+1) + " - Damage done: " + str(players[i].calculate_damage()))

        damage_taken = players[i].damaged(100)
        if damage_taken[0] == True:
            print("Player " + str(i+1) + " - Damage taken for 100: " + str(damage_taken[1]) + "\n")
        else:
            print("Player " + str(i+1) + " - Dodged!" + "\n")

    # Actual game starts here: ------------
    
    for turns in range(0, 150):
        print("\n> Starting turn " + str(turns))

        # Enemy attacks
        raw_dmg = Enemy.calculate_damage()
        target = random.randint(0, len(player_count) - 1)
        damage_taken = players[target].damaged(raw_dmg)
        print(Enemy.name + " attacks! Throws " + str(raw_dmg) + " dmg at player " + str(target))
        print("Player " + str(target) + " takes " + str(damage_taken) + "!")

        # Players attack

wb.close() # Close the doc to avoid corruption

if __name__ == "__main__":
    main()
