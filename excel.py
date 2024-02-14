import openpyxl as pxl
import random
import ast

# Open and setup doc
wb = pxl.load_workbook('doc.xlsx', data_only=True)
sheets = wb.sheetnames

def read_file(filename, index):
    with open(filename, 'r') as file:
        lines = file.readlines() # Read the file
        if index < len(lines):
            return lines[index].strip()
        else:
            return None  # Index out of range

def read_excel(sheet, cell): # Read the value on a cell, in a specified sheet of the doc
    sheetName = str(sheet)
    ws = wb[sheetName]
    return ws[str(cell)].value

class Weapon:
    def __init__(self, name, rolls, hit_threshold):
        self.name = name
        self.rolls = int(rolls)
        self.hit_threshold = int(hit_threshold)
    
    def display(self):
        return self.name, self.rolls, self.hit_threshold

class Item:
    def __init__(self, name, defense, dodge, crit):
        self.name = name
        self.defense = defense
        self.dodge = dodge
        self.crit = crit
    
    def display(self):
        return self.name, self.defense, self.dodge, self.crit

class Player:
    def __init__(self, level, role, weapon_ID, item_ID):
        self.role = str(role)
        self.level = str(level)
        self.weapon = Weapon( # Using the Weapon class in the Player class here
            name = read_excel("Weapons", str("C" + str(int(weapon_ID) + 1))),
            rolls = read_excel("Weapons", str("D" + str(int(weapon_ID) + 1))),
            hit_threshold = read_excel("Weapons", str("E" + str(int(weapon_ID) + 1)))
        )
        self.item = Item(
            name = read_excel("Items", str("C" + str(int(item_ID) + 1))),
            defense = read_excel("Items", str("D" + str(int(item_ID) + 1))),
            dodge = read_excel("Items", str("E" + str(int(item_ID) + 1))),
            crit = read_excel("Items", str("F" + str(int(item_ID) + 1)))
        )

    def calculate_damage_rolls(self):
        damage = 0
        for _ in range(self.weapon.rolls):
            roll_result = random.randint(1, 6)  # Roll a dice
            if roll_result >= self.weapon.hit_threshold:
                damage += 1
        
        if random.randint(1, 100) <= int(self.item.crit):
            return damage*2
        else:
            return damage

    def damaged(self, dmg):
        roll = random.randint(1, 100)
        if roll >= int(self.item.dodge):
            return True, dmg - int(self.item.defense)/100*dmg
        return False, 0
        
    def display(self):
        return self.role, self.level, self.weapon.display(), self.item.display()

def main():
    # Values defined by a tk interface, written in a config.txt file
    config = ast.literal_eval(read_file("config.txt", 0)) # Convert the config file's info to array
    player_1_role = config[0][0]
    player_1_level = config[0][1]
    player_1_weapon_ID = config[0][2]
    player_1_item_ID = config[0][3]

    # Setup from the doc:
    player_1 = Player( # To be set in a loop for more players
        role = player_1_role,
        level = player_1_level,
        weapon_ID = player_1_weapon_ID,
        item_ID = player_1_item_ID
    )

    # Then temporary, just debugging for now
    print(player_1.display())

    print("Damage done: " + str(player_1.calculate_damage_rolls()))

    damage_taken = player_1.damaged(100)
    if damage_taken[0] == True:
        print("Damage taken for 100: " + str(damage_taken[1]))
    else:
        print("Dodged!")

wb.close() # Close the doc to avoid corruption

if __name__ == "__main__":
    main()
